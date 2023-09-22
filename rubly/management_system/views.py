from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.db.models import F
from openpyxl import Workbook
import openpyxl
import urllib
import ast
import io
import zipfile
from openpyxl import Workbook
from io import BytesIO
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated 
from .serializers import GoodsReceivedSerializer
from django.contrib import messages, auth
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.shortcuts import render, redirect
from django.views import View
import json
from django.contrib.auth.models import User
from django.http import JsonResponse
import os
import mimetypes
from django.core.serializers import serialize
import time
from django.db.models import Sum
import datetime
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import api_view
from openpyxl.styles import PatternFill, Font
import base64
def Login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = auth.authenticate(username=username, password=password)
        if user is not None:
            auth.login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid Credentials')
            return redirect('Login')
    else:
        return render(request, 'auth-login-basic.html')


# Dashboard Module
@login_required(login_url="Login")
def Dashboard(request):
        project_types=Project_Type.objects.all()
        clients=Client.objects.all()
        
        list_clients=[]
        for client in clients:
            projects=[]
            for project in project_types:
                if project.client.name==client.name:
                    Project_region = Project.objects.filter(project_type__name=project.name,project_type__client__name=client.name)
                    projects.append({"project_type":project,"project_region":Project_region})

            list_clients.append({
                "client":client,
                "projects":projects,
            })
        Pur=Purchase_Order.objects.all()
        POs=[]
        for p in Pur:
            if p.purchase_ID not in POs:
                POs.append(p.purchase_ID)
        issuance=IssuanceExternal.objects.all()
        companies=[]
        for iss in issuance:
            if iss.company not in companies:
                companies.append(iss.company)
        instances=[]
        for comp in companies:
            val=IssuanceExternal.objects.filter(company=comp).count()
            instances.append({"company":comp,"frequency":val})
        sorted_list = sorted(instances, key=lambda x: x['frequency'], reverse=True)
        return render(request, 'dashboard.html',{"list_clients":list_clients,"purchase_orders":POs,"objects":sorted_list})
@login_required(login_url="Login")
def issue(request):
    clients=Client.objects.all()
    project_types=Project_Type.objects.all()
    objects=[]
    for clien in clients:
        project=[]
        for pro in project_types:
            if pro.client.name==clien.name:
                project.append(pro.name)
        objects.append({"clients":clien,"projects":project})
    return render(request,"issue.html",{"objects":objects})
@login_required(login_url="Login")
def check(request):
    if request.GET['client']:
        client=request.GET['client']
        proje=request.GET['project']
        project=Project_Type.objects.get(client__name=client,name=proje)
        #get all goods where the project type is the one selected
        pos=Purchase_Order.objects.filter(project_type=project)
        #get all goods where the project type is the one selected
        all_goods=Goods_received.objects.all()
        goods=[]
        for po in pos:
            for good in all_goods:
                if good not in goods and good.Purchase_Order == po:
                    goods.append(good)
        des=[]
        #looping through the goods and adding their description in the list
        for good in goods:
            if good.description not in des:
                des.append(good.description)
        #loop on every descriptio appending the description instance and all the goods associated with it
        objects=[]
        for description in des:
            #loop on every good testing if the good.description matches the one for our current loop then adding it in the list
            Quantity=0
            for good in goods:
                if good.description==description:
                    Quantity+=good.remaining
            objects.append({"description":description,"quantity":Quantity})
        if objects==[]:
            message="No goods for selected type"
        else:
            message="found"
    return JsonResponse(message,safe=False)
@login_required(login_url="Login")
def projectGoods(request):
    if request.GET['client']:
        client=request.GET['client']
        proje=request.GET['project']
        project=Project_Type.objects.get(client__name=client,name=proje)
        pos=Purchase_Order.objects.filter(project_type=project)
        #get all goods where the project type is the one selected
        all_goods=Goods_received.objects.all()
        goods=[]
        for po in pos:
            for good in all_goods:
                if good not in goods and good.Purchase_Order == po:
                    goods.append(good)
        des=[]
        #looping through the goods and adding their description in the list
        for good in goods:
            if good.description not in des:
                des.append(good.description)
        #loop on every descriptio appending the description instance and all the goods associated with it
        objects=[]
        for description in des:
            instan=Description.objects.get(Description=description)
            #loop on every good testing if the good.description matches the one for our current loop then adding it in the list
            Quantity=0
            for good in goods:
                if good.description==description:
                    Quantity+=good.remaining
            objects.append({"description":instan,"quantity":Quantity})
        return render(request,"projectGoods.html",{"objects":objects,"client":client,"project":proje})
@login_required(login_url="Login")
def good(request):
    if request.GET['description']:
        des=request.GET['description']
        client=request.GET['client']
        proj=request.GET['project']
        goods=Goods_received.objects.filter(Purchase_Order__project_type__name=proj,description__Description=des,remaining__gt=0)
        desc=Description.objects.get(Description=des)
        Quantity=0
        for good in goods:
                Quantity+=good.remaining
        goods_json = []
        for good in goods:
            goods_json.append({"po":good.Purchase_Order.purchase_ID,"quantity":good.remaining})
        return render(request,"good.html",{"goods":goods,"description":desc,"quantity":Quantity,"client":client,"project":proj,"goods_json":goods_json})
@login_required(login_url="Login")
def checkout(request):
    now=datetime.datetime.now()
    date=now.strftime("%Y-%m-%d")
    if request.GET['company']:
        po=Purchase_Order.objects.get(purchase_ID=request.GET['po'],project_type__client__name=request.GET['client'])
        good=Goods_received.objects.get(Purchase_Order__purchase_ID=request.GET['po'],description__Description=request.GET['desc'])
        good.remaining=good.remaining-int(request.GET['quantity'])
        material=Description.objects.get(Description=request.GET['desc'])
        mgood=Goods_received.objects.get(Purchase_Order__purchase_ID=request.GET['po'],description=material,Purchase_Order=po)
        issue=IssuanceExternal(date=date,good=mgood,company=request.GET['company'],Carpex=request.GET['capex'],Quantity=request.GET['quantity'],remaining=good.remaining)
        try:
            good.save()
            issue.save()
            return JsonResponse({"message":"Successfull"})
        except:
            return JsonResponse({"message":"An error occured"})
@login_required(login_url="Login")
def capex(request):
    download=request.GET.get('download','')
    if download:
        client_id=request.GET.get('client_id','')
        po_id=request.GET.get('po_id','')
        clients=[]
        pos=[]
        issued=IssuanceExternal.objects.all()
        if po_id:
            pos.append(Purchase_Order.objects.get(id=po_id))
            for iss in issued:
                recieved=iss.good
                if recieved.Purchase_Order.project_type.client not in clients:
                    clients.append(recieved.Purchase_Order.project_type.client)
        if client_id:
            clients.append(Client.objects.get(id=client_id))
            for iss in issued:
                recieved=iss.good
                if recieved.Purchase_Order not in pos:
                    pos.append(recieved.Purchase_Order)
        elif "po_id" not in request.GET and "client_id" not in request.GET:
            for iss in issued:
                recieved=iss.good
                if recieved.Purchase_Order.project_type.client not in clients:
                    clients.append(recieved.Purchase_Order.project_type.client)
                if recieved.Purchase_Order not in pos:
                    pos.append(recieved.Purchase_Order)
        types=[]
        desc=[]
        goods=[]
        for iss in issued:
            try:
                recieved=iss.good
            except:
                return render(request,'capex.html',{"message":"There was an error in fetching the capexex"})
            if recieved not in goods:
                goods.append(recieved)
            if iss.good.description.Type not in types:
                types.append(iss.good.description.Type)
            if iss.good.description not in desc:
                desc.append(iss.good.description)
        capexes=IssuanceExternal.objects.all()
        objects=[]
        for client in clients:
            ob_pos=[]
            for po in pos:
                ob_types=[]
                for stype in types:
                    ob_goods=[]
                    for good in goods:
                        if good.Purchase_Order.project_type.client == client and good.description.Type == stype and good.Purchase_Order == po:
                            ob_capexex=[]
                            for capex in capexes:
                                if capex.good == good and capex not in ob_capexex:
                                    ob_capexex.append(capex)
                            if ob_capexex !=[]:
                                sorted_capexex = sorted(ob_capexex, key=lambda c: c.remaining, reverse=True)
                                ob_goods.append({"good":good,"capex":sorted_capexex})
                    if ob_goods !=[]:
                        ob_types.append({"type":stype,"ob_goods":ob_goods})
                if ob_types !=[]:
                    ob_pos.append({"po":po,"ob_types":ob_types})
            if ob_pos !=[]:
                objects.append({"client":client,"ob_pos":ob_pos})
        
        for c in objects:
             name=c['client'].name+" "+ "Capexes"
             row_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
             fill = PatternFill(start_color="7FB3D5", end_color="7FB3D5", fill_type="solid")
             workbook=openpyxl.Workbook()
             count=0
             for po in c['ob_pos']:
                if count==0:
                    default_sheet = workbook.active
                    default_sheet.title = po['po'].purchase_ID
                    worksheet = default_sheet
                else:
                    worksheet = workbook.create_sheet(title=po['po'].purchase_ID)
                worksheet['A1']="TYPE"
                worksheet['A1'].fill=row_fill
                worksheet['B1']="MATERIAL"
                worksheet['B1'].fill=row_fill
                worksheet['C1']="PACKAGING"
                worksheet['C1'].fill=row_fill
                worksheet['D1']="RECEIVED"
                worksheet['D1'].fill=row_fill
                worksheet['E1']="CAPEX"
                worksheet['E1'].fill=row_fill
                worksheet['F1']="COMPANY"
                worksheet['F1'].fill=row_fill
                worksheet['G1']="QUANTITY"
                worksheet['G1'].fill=row_fill
                worksheet['H1']="REMAINING"
                worksheet['H1'].fill=row_fill
                bold_range = worksheet['A1:H1']

                # Create a Font object with bold set to True
                bold_font = Font(bold=True)

                # Apply the bold font to the specified range of cells
                for row in bold_range:
                    for cell in row:
                        cell.font = bold_font
                # Auto-adjust column widths to fit the content
                for type in po['ob_types']:
                    count1=0
                    for good in type['ob_goods']:
                        count2=0
                        for cap in good['capex']:
                            print("count1:",count1)
                            print("count1:",count2)
                            if count1==0:
                                row_data=[type['type'].name,good['good'].description.Description,good['good'].description.Packaging,good['good'].Quantity,cap.Carpex,cap.company,cap.Quantity,cap.remaining]
                                next_row = worksheet.max_row + 1
                                worksheet.append(row_data)
                                # Append the row data to the worksheet in the next available row
                                for col_num, value in enumerate(row_data, start=1):
                                    worksheet.cell(row=next_row, column=col_num, value=value).fill = fill
                                count2+=1
                                count1+=1
                                continue
                        
                            else:
                                if count2==0:
                                    print("good1")
                                    row_data=["--",good['good'].description.Description,good['good'].description.Packaging,good['good'].Quantity,cap.Carpex,cap.company,cap.Quantity,cap.remaining]
                                    worksheet.append(row_data)
                                    count2+=1
                                    count1+=1
                                    continue
                                else:
                                    print("other")
                                    row_data=["--","--","--",'--',cap.Carpex,cap.company,cap.Quantity,cap.remaining]
                                    worksheet.append(row_data)
                                    count2+=1
                                    count1+=1
                                    continue
                            
                
                for column_cells in worksheet.columns:
                    max_length = 0
                    column = column_cells[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C', ...)
                    
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass

                    adjusted_width = (max_length + 2)  # Add a little padding
                    worksheet.column_dimensions[column].width = adjusted_width
                count+=1
             response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
             response['Content-Disposition'] = f'attachment; filename={name}.xlsx'
             workbook.save(response)
             return response
    client_id=request.GET.get('client_id','')
    po_id=request.GET.get('po_id','')
    clients=[]
    pos=[]
    issued=IssuanceExternal.objects.all()
    if po_id:
        pos.append(Purchase_Order.objects.get(id=po_id))
        for iss in issued:
            recieved=iss.good
            if recieved.Purchase_Order.project_type.client not in clients:
                clients.append(recieved.Purchase_Order.project_type.client)
    if client_id:
        clients.append(Client.objects.get(id=client_id))
        for iss in issued:
            recieved=iss.good
            if recieved.Purchase_Order not in pos:
                pos.append(recieved.Purchase_Order)
    elif "po_id" not in request.GET and "client_id" not in request.GET:
        for iss in issued:
            recieved=iss.good
            if recieved.Purchase_Order.project_type.client not in clients:
                clients.append(recieved.Purchase_Order.project_type.client)
            if recieved.Purchase_Order not in pos:
                pos.append(recieved.Purchase_Order)
    types=[]
    desc=[]
    goods=[]
    for iss in issued:
        try:
            recieved=iss.good
        except:
            return render(request,'capex.html',{"message":"There was an error in fetching the capexex"})
        if recieved not in goods:
            goods.append(recieved)
        if iss.good.description.Type not in types:
            types.append(iss.good.description.Type)
        if iss.good.description not in desc:
            desc.append(iss.good.description)
    capexes=IssuanceExternal.objects.all()
    objects=[]
    for client in clients:
        ob_pos=[]
        for po in pos:
            ob_types=[]
            for stype in types:
                ob_goods=[]
                for good in goods:
                    if good.Purchase_Order.project_type.client == client and good.description.Type == stype and good.Purchase_Order == po:
                        ob_capexex=[]
                        for capex in capexes:
                            if capex.good == good and capex not in ob_capexex:
                                ob_capexex.append(capex)
                        if ob_capexex !=[]:
                            sorted_capexex = sorted(ob_capexex, key=lambda c: c.remaining, reverse=True)
                            ob_goods.append({"good":good,"capex":sorted_capexex})
                if ob_goods !=[]:
                    ob_types.append({"type":stype,"ob_goods":ob_goods})
            if ob_types !=[]:
                ob_pos.append({"po":po,"ob_types":ob_types})
        if ob_pos !=[]:
            objects.append({"client":client,"ob_pos":ob_pos})
    return render(request,'capex.html',{"objects":objects})
def Logout(request):
    if request.method == 'POST':
        auth.logout(request)
        return redirect('Login')
    return redirect('Login')

# Dashboard-date Module
# @login_required(login_url="Login")
# def stock(request):
#     if request.GET['selected']=="all":
#          clients=Client.objects.all()
#          recieved=Goods_received.objects.all()
#          return render(request,"stock.html")
#     else:
#         clients=Client.objects.all()

#         return render(request,"stock.html")
@login_required(login_url="Login")
def dashboardstock(request):
    if request.method=="GET":
        if 'openingdate' in request.GET:
            respdict=[]
            date=request.GET['openingdate']
            search = 'opening' + date + '.json'
            file_path = os.path.join("data", f"{search}")
            with open(file_path, 'r') as file:
                datav = json.load(file)
                for jsonval in datav :
                    sublist=[]
                    for i in range(len(jsonval)):
                        if i==0:
                            sublist.append({"Description":jsonval[i]})
                        if i==1:
                            sublist.append({"Price":jsonval[i]})
                        if i==2:
                            sublist.append({"Packaging":jsonval[i]})
                        if i==3:
                            sublist.append({"Quantity":jsonval[i]})
                        if i==4:
                            sublist.append({"Purchase_order":jsonval[i]})
                    respdict.append(sublist)
                text="Opening Stock For:"+str(date)
            return render(request, 'dashboard.html',{'respdict': respdict,"type":text})
        elif 'closingdate' in request.GET:
            respdict=[]
            date=request.GET['closingdate']
            search = 'closing' + date + '.json'
            file_path = os.path.join("data", f"{search}")
            with open(file_path, 'r') as file:
                datav = json.load(file)
                for jsonval in datav :
                    sublist=[]
                    for i in range(len(jsonval)):
                        if i==0:
                            sublist.append({"Description":jsonval[i]})
                        if i==1:
                            sublist.append({"Price":jsonval[i]})
                        if i==2:
                            sublist.append({"Packaging":jsonval[i]})
                        if i==3:
                            sublist.append({"Quantity":jsonval[i]})
                        if i==4:
                            sublist.append({"Purchase_order":jsonval[i]})
                    respdict.append(sublist)
            text="Closing Stock for:"+str(date)
            return render(request, 'dashboard.html',{'respdict': respdict,"type":text})
        


def current_stocks_list(request):
    download=request.GET.get('download','')
    if download:
        client_id = request.GET.get('client_id')
        type_id=request.GET.get('type_id')
        allGoods=Goods_received.objects.filter(remaining__gt=0)
        types=[]
        des=[]
        clients=[]
        if client_id:
            clients.append(Client.objects.get(id=client_id))
            for good in allGoods:
                if good.description.Type not in types:
                    types.append(good.description.Type)
                if good.description not in des:
                    des.append(good.description)
        if type_id:
            types.append(Description_Type.objects.get(id=type_id))
            for good in allGoods:
                    if good.Purchase_Order.project_type.client not in clients:
                        clients.append(good.Purchase_Order.project_type.client)
                    if good.description not in des:
                        des.append(good.description)
        elif "client_id" not in request.GET and 'type_id' not in request.GET:
            for good in allGoods:
                if good.Purchase_Order.project_type.client not in clients:
                    clients.append(good.Purchase_Order.project_type.client)
                if good.description.Type not in types:
                    types.append(good.description.Type)
                if good.description not in des:
                    des.append(good.description)
        ob_clients=[]
        for c in clients:
            ob_types=[]
            for t in types:
                ob_desc=[]
                for d in des:
                    if d.Type == t:
                        ob_goods=[]
                        total=0
                        for good in allGoods:
                            if good.Purchase_Order.project_type.client == c and good.description == d:
                                total+=good.remaining
                                try:
                                    ob_goods.append({"good":good,"price":int(good.price)*good.remaining})
                                except:
                                        ob_goods.append({"good":good,"price":"None"})
                        if ob_goods !=[]:
                            ob_desc.append({"desc":d,"goods":ob_goods,"total":total})
                if ob_desc !=[]:
                    ob_types.append({"type":t,"desc":ob_desc})
            if ob_types !=[]:
                ob_clients.append({"client":c,"types":ob_types})
        workbook = openpyxl.Workbook()
        row_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
        fill = PatternFill(start_color="7FB3D5", end_color="7FB3D5", fill_type="solid")
        count =0
        for c in ob_clients:
             if count==0:
                default_sheet = workbook.active
                default_sheet.title = c['client'].name
                worksheet = default_sheet
             else:
                worksheet = workbook.create_sheet(title=c['client'].name)
             worksheet['A1']="TYPE"
             worksheet['A1'].fill=row_fill
             worksheet['B1']="MATERIAL"
             worksheet['B1'].fill=row_fill
             worksheet['C1']="PACKAGING"
             worksheet['C1'].fill=row_fill
             worksheet['D1']="Total"
             worksheet['D1'].fill=row_fill
             worksheet['E1']="PO"
             worksheet['E1'].fill=row_fill
             worksheet['F1']="Quantity"
             worksheet['F1'].fill=row_fill
             worksheet['G1']="Project Type"
             worksheet['G1'].fill=row_fill
            
             bold_range = worksheet['A1:G1']

             # Create a Font object with bold set to True
             bold_font = Font(bold=True)

             # Apply the bold font to the specified range of cells
             for row in bold_range:
                for cell in row:
                    cell.font = bold_font
            # Auto-adjust column widths to fit the content
             
             
             for t in c['types']:
                 count1=0
                 for m in t['desc']:
                     count2=0
                     for g in m['goods']:
                        if count1==0:

                            row_data=[t['type'].name,m['desc'].Description,m['desc'].Packaging,m['total'],g['good'].Purchase_Order.purchase_ID,g['good'].remaining,g['good'].Purchase_Order.project_type.name]
                            next_row = worksheet.max_row + 1

                            # Append the row data to the worksheet in the next available row
                            for col_num, value in enumerate(row_data, start=1):
                                worksheet.cell(row=next_row, column=col_num, value=value).fill = fill
                            count2+=1
                            count1+=1
                            continue
                    
                        else:
                            if count2==0:
                                row_data=['--',m['desc'].Description,m['desc'].Packaging,m['total'],g['good'].Purchase_Order.purchase_ID,g['good'].remaining,g['good'].Purchase_Order.project_type.name]
                                worksheet.append(row_data)
                                count2+=1
                                count1+=1
                                continue
                            else:
                                row_data=["--",'--','--','--',g['good'].Purchase_Order.purchase_ID,g['good'].remaining,g['good'].Purchase_Order.project_type.name]
                                worksheet.append(row_data)
                                count2+=1
                                count1+=1
                                continue
                     count+=1
             for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C', ...)
                
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                adjusted_width = (max_length + 2)  # Add a little padding
                worksheet.column_dimensions[column].width = adjusted_width
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=stocks.xlsx'
        workbook.save(response)
        return response

    client_id = request.GET.get('client_id')
    type_id=request.GET.get('type_id')
    allGoods=Goods_received.objects.filter(remaining__gt=0)
    types=[]
    des=[]
    clients=[]
    if client_id:
        clients.append(Client.objects.get(id=client_id))
        for good in allGoods:
            if good.description.Type not in types:
                types.append(good.description.Type)
            if good.description not in des:
                des.append(good.description)
    if type_id:
       types.append(Description_Type.objects.get(id=type_id))
       for good in allGoods:
            if good.Purchase_Order.project_type.client not in clients:
                clients.append(good.Purchase_Order.project_type.client)
            if good.description not in des:
                des.append(good.description)
    elif "client_id" not in request.GET and 'type_id' not in request.GET:
        for good in allGoods:
            if good.Purchase_Order.project_type.client not in clients:
                clients.append(good.Purchase_Order.project_type.client)
            if good.description.Type not in types:
                types.append(good.description.Type)
            if good.description not in des:
                des.append(good.description)
    ob_clients=[]
    for c in clients:
        ob_types=[]
        for t in types:
            ob_desc=[]
            for d in des:
                if d.Type == t:
                    ob_goods=[]
                    total=0
                    for good in allGoods:
                        if good.Purchase_Order.project_type.client == c and good.description == d:
                            total+=good.remaining
                            try:
                                ob_goods.append({"good":good,"price":int(good.price)*good.remaining})
                            except:
                                    ob_goods.append({"good":good,"price":"None"})
                    if ob_goods !=[]:
                        ob_desc.append({"desc":d,"goods":ob_goods,"total":total})
            if ob_desc !=[]:
                ob_types.append({"type":t,"desc":ob_desc})
        if ob_types !=[]:
            ob_clients.append({"client":c,"types":ob_types})
    return render(request, 'stock.html',{"objects":ob_clients})
def po_stock(request):
    downloadsingle=request.GET.get('downloadsingle','')
    download=request.GET.get('download','')
    if downloadsingle:
            po=request.GET.get("po_id")
            client_id=request.GET.get("client_id")
            proj_id=request.GET.get("proj_id")
            allGoods=Goods_received.objects.filter(remaining__gt=0)
            clients=[]
            proj_types=[]
            pos=[]
            types=[]
            if po:
                pos.append(Purchase_Order.objects.get(id=po))
                print(pos)
                des=[]
                for good in allGoods:
                    if good.Purchase_Order.project_type not in proj_types:
                        proj_types.append(good.Purchase_Order.project_type)
                    if good.Purchase_Order.project_type.client not in clients:
                        clients.append(good.Purchase_Order.project_type.client)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            if client_id:
                des=[]
                clients.append(Client.objects.get(id=client_id))
                for good in allGoods:
                    if good.Purchase_Order.project_type not in proj_types:
                        proj_types.append(good.Purchase_Order.project_type)
                    if good.Purchase_Order not in pos:
                        pos.append(good.Purchase_Order)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            if proj_id:
                proj_types.append(Project_Type.objects.get(id=proj_id))
                des=[]
                for good in allGoods:
                    if good.Purchase_Order not in pos:
                        pos.append(good.Purchase_Order)
                    if good.Purchase_Order.project_type.client not in clients:
                        clients.append(good.Purchase_Order.project_type.client)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            elif "po_id" not in request.GET and "proj_id" not in request.GET and 'client_id' not in request.GET:
                des=[]
                for good in allGoods:
                    if good.Purchase_Order.project_type not in proj_types:
                        proj_types.append(good.Purchase_Order.project_type)
                    if good.Purchase_Order not in pos:
                        pos.append(good.Purchase_Order)
                    if good.Purchase_Order.project_type.client not in clients:
                        clients.append(good.Purchase_Order.project_type.client)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            ob_clients=[]
            for c in clients:
                ob_projects=[]
                for proj in proj_types:
                    if proj.client == c:
                        ob_pos =[]
                        for po in pos:
                            if po.project_type == proj:
                                ob_types=[]
                                for t in types:
                                    ob_goods =[]
                                    for good in allGoods:
                                        if good.Purchase_Order == po and good.description.Type == t:
                                            ob_goods.append(good)
                                    if ob_goods!=[]:
                                        ob_types.append({"type":t,"goods":ob_goods})
                                if ob_types !=[]:
                                    ob_pos.append({"po":po,"types":ob_types})
                        if ob_pos != []:
                            ob_projects.append({"project":proj,"pos":ob_pos}) 
                if ob_projects != []:
                    ob_clients.append({"client":c,"projects":ob_projects})
            for stock in ob_clients:
                    workbook = Workbook()
                    count= 0
                    name=stock['client'].name
                    for proj in stock['projects']:
                        for po in proj['pos']:
                            if count==0:
                                default_sheet = workbook.active
                                default_sheet.title = po['po'].purchase_ID
                                sheet = default_sheet
                                count+=1
                            else:
                                sheet = workbook.create_sheet(title=po['po'].purchase_ID)
                                count+=1
                            row_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
                            fill = PatternFill(start_color="F2D7D5", end_color="F2D7D5", fill_type="solid")
                            sheet['A1']="TYPE"
                            sheet['A1'].fill=row_fill
                            sheet['B1']="DESCRIPTION"
                            sheet['B1'].fill=row_fill
                            sheet['C1']="PACKAGING"
                            sheet['C1'].fill=row_fill
                            sheet['D1']="QUANTITY"
                            sheet['D1'].fill=row_fill
                            sheet['E1']="UNIT PRICE"
                            sheet['E1'].fill=row_fill
                            bold_range = sheet['A1:E1']

                            # Create a Font object with bold set to True
                            bold_font = Font(bold=True)

                            # Apply the bold font to the specified range of cells
                            for row in bold_range:
                                for cell in row:
                                    cell.font = bold_font
                            for typ in po['types']:
                                count1=0
                                for good in typ['goods']:
                                    if good.price == None:
                                        price = "Null"
                                    else:
                                        price= good.price
                                    if count1==0:
                                        row_data=[typ['type'].name,good.description.Description,good.description.Packaging,good.remaining,price]
                                        next_row = sheet.max_row + 1

                                        # Append the row data to the worksheet in the next available row
                                        for col_num, value in enumerate(row_data, start=1):
                                            sheet.cell(row=next_row, column=col_num, value=value).fill = fill
                                        count1+=1
                                        continue
                                    else:
                                        row_data=["--",good.description.Description,good.description.Packaging,good.remaining,price]
                                        sheet.append(row_data)
                                        count1+=1
                            for column_cells in sheet.columns:
                                max_length = 0
                                column = column_cells[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C', ...)
                                
                                for cell in column_cells:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)  # Add a little padding
                                sheet.column_dimensions[column].width = adjusted_width
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename={name}.xlsx'
                    workbook.save(response)
                    return response
    if download:
           
            po=request.GET.get("po_id")
            client_id=request.GET.get("client_id")
            proj_id=request.GET.get("proj_id")
            allGoods=Goods_received.objects.filter(remaining__gt=0)
            clients=[]
            proj_types=[]
            pos=[]
            types=[]
            if po:
                pos.append(Purchase_Order.objects.get(id=po))
                print(pos)
                des=[]
                for good in allGoods:
                    if good.Purchase_Order.project_type not in proj_types:
                        proj_types.append(good.Purchase_Order.project_type)
                    if good.Purchase_Order.project_type.client not in clients:
                        clients.append(good.Purchase_Order.project_type.client)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            if client_id:
                des=[]
                clients.append(Client.objects.get(id=client_id))
                for good in allGoods:
                    if good.Purchase_Order.project_type not in proj_types:
                        proj_types.append(good.Purchase_Order.project_type)
                    if good.Purchase_Order not in pos:
                        pos.append(good.Purchase_Order)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            if proj_id:
                proj_types.append(Project_Type.objects.get(id=proj_id))
                des=[]
                for good in allGoods:
                    if good.Purchase_Order not in pos:
                        pos.append(good.Purchase_Order)
                    if good.Purchase_Order.project_type.client not in clients:
                        clients.append(good.Purchase_Order.project_type.client)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            elif "po_id" not in request.GET and "proj_id" not in request.GET and 'client_id' not in request.GET:
                des=[]
                for good in allGoods:
                    if good.Purchase_Order.project_type not in proj_types:
                        proj_types.append(good.Purchase_Order.project_type)
                    if good.Purchase_Order not in pos:
                        pos.append(good.Purchase_Order)
                    if good.Purchase_Order.project_type.client not in clients:
                        clients.append(good.Purchase_Order.project_type.client)
                    if good.description.Type not in types:
                        types.append(good.description.Type)
                    if good.description not in des:
                        des.append(good.description)
            ob_clients=[]
            for c in clients:
                ob_projects=[]
                for proj in proj_types:
                    if proj.client == c:
                        ob_pos =[]
                        for po in pos:
                            if po.project_type == proj:
                                ob_types=[]
                                for t in types:
                                    ob_goods =[]
                                    for good in allGoods:
                                        if good.Purchase_Order == po and good.description.Type == t:
                                            ob_goods.append(good)
                                    if ob_goods!=[]:
                                        ob_types.append({"type":t,"goods":ob_goods})
                                if ob_types !=[]:
                                    ob_pos.append({"po":po,"types":ob_types})
                        if ob_pos != []:
                            ob_projects.append({"project":proj,"pos":ob_pos}) 
                if ob_projects != []:
                    ob_clients.append({"client":c,"projects":ob_projects}) 
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, False) as zipf:
        
                # excel_files = []

                for stock in ob_clients:
                    workbook = Workbook()
                    count= 0
                    for proj in stock['projects']:
                        for po in proj['pos']:
                            if count==0:
                                default_sheet = workbook.active
                                default_sheet.title = po['po'].purchase_ID
                                sheet = default_sheet
                                count+=1
                            else:
                                sheet = workbook.create_sheet(title=po['po'].purchase_ID)
                                count+=1
                            row_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
                            fill = PatternFill(start_color="F2D7D5", end_color="F2D7D5", fill_type="solid")
                            sheet['A1']="TYPE"
                            sheet['A1'].fill=row_fill
                            sheet['B1']="DESCRIPTION"
                            sheet['B1'].fill=row_fill
                            sheet['C1']="PACKAGING"
                            sheet['C1'].fill=row_fill
                            sheet['D1']="QUANTITY"
                            sheet['D1'].fill=row_fill
                            sheet['E1']="UNIT PRICE"
                            sheet['E1'].fill=row_fill
                            bold_range = sheet['A1:E1']

                            # Create a Font object with bold set to True
                            bold_font = Font(bold=True)

                            # Apply the bold font to the specified range of cells
                            for row in bold_range:
                                for cell in row:
                                    cell.font = bold_font
                            for typ in po['types']:
                                count1=0
                                for good in typ['goods']:
                                    if good.price == None:
                                        price = "Null"
                                    else:
                                        price= good.price
                                    if count1==0:
                                        row_data=[typ['type'].name,good.description.Description,good.description.Packaging,good.remaining,price]
                                        next_row = sheet.max_row + 1

                                        # Append the row data to the worksheet in the next available row
                                        for col_num, value in enumerate(row_data, start=1):
                                            sheet.cell(row=next_row, column=col_num, value=value).fill = fill
                                        count1+=1
                                        continue
                                    else:
                                        row_data=["--",good.description.Description,good.description.Packaging,good.remaining,price]
                                        sheet.append(row_data)
                                        count1+=1
                            for column_cells in sheet.columns:
                                max_length = 0
                                column = column_cells[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C', ...)
                                
                                for cell in column_cells:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)  # Add a little padding
                                sheet.column_dimensions[column].width = adjusted_width
                    # Create an empty zip file
                    

                    

                    excel_buffer = BytesIO()
                    workbook.save(excel_buffer)

                    # Add the Excel data to the zip file with a unique name
                    zipf.writestr(f'{stock["client"].name}.xlsx', excel_buffer.getvalue())

            # Set up the HTTP response with the zip file
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=stocks.zip'
            response['Content-Length'] = zip_buffer.tell()
            zip_buffer.seek(0)

            # Return the zip file as the response
            return response





    po=request.GET.get("po_id")
    client_id=request.GET.get("client_id")
    proj_id=request.GET.get("proj_id")
    allGoods=Goods_received.objects.filter(remaining__gt=0)
    clients=[]
    proj_types=[]
    pos=[]
    types=[]
    if po:
        pos.append(Purchase_Order.objects.get(id=po))
        print(pos)
        des=[]
        for good in allGoods:
            if good.Purchase_Order.project_type not in proj_types:
                proj_types.append(good.Purchase_Order.project_type)
            if good.Purchase_Order.project_type.client not in clients:
                clients.append(good.Purchase_Order.project_type.client)
            if good.description.Type not in types:
                types.append(good.description.Type)
            if good.description not in des:
                des.append(good.description)
    if client_id:
        des=[]
        clients.append(Client.objects.get(id=client_id))
        for good in allGoods:
            if good.Purchase_Order.project_type not in proj_types:
                proj_types.append(good.Purchase_Order.project_type)
            if good.Purchase_Order not in pos:
                pos.append(good.Purchase_Order)
            if good.description.Type not in types:
                types.append(good.description.Type)
            if good.description not in des:
                des.append(good.description)
    if proj_id:
        proj_types.append(Project_Type.objects.get(id=proj_id))
        des=[]
        for good in allGoods:
            if good.Purchase_Order not in pos:
                pos.append(good.Purchase_Order)
            if good.Purchase_Order.project_type.client not in clients:
                clients.append(good.Purchase_Order.project_type.client)
            if good.description.Type not in types:
                types.append(good.description.Type)
            if good.description not in des:
                des.append(good.description)
    elif "po_id" not in request.GET and "proj_id" not in request.GET and 'client_id' not in request.GET:
        des=[]
        for good in allGoods:
            if good.Purchase_Order.project_type not in proj_types:
                proj_types.append(good.Purchase_Order.project_type)
            if good.Purchase_Order not in pos:
                pos.append(good.Purchase_Order)
            if good.Purchase_Order.project_type.client not in clients:
                clients.append(good.Purchase_Order.project_type.client)
            if good.description.Type not in types:
                types.append(good.description.Type)
            if good.description not in des:
                des.append(good.description)
    ob_clients=[]
    for c in clients:
        ob_projects=[]
        for proj in proj_types:
            if proj.client == c:
                ob_pos =[]
                for po in pos:
                    if po.project_type == proj:
                        ob_types=[]
                        for t in types:
                            ob_goods =[]
                            for good in allGoods:
                                if good.Purchase_Order == po and good.description.Type == t:
                                    ob_goods.append(good)
                            if ob_goods!=[]:
                                ob_types.append({"type":t,"goods":ob_goods})
                        if ob_types !=[]:
                            ob_pos.append({"po":po,"types":ob_types})
                if ob_pos != []:
                    ob_projects.append({"project":proj,"pos":ob_pos}) 
        if ob_projects != []:
            ob_clients.append({"client":c,"projects":ob_projects}) 
    return render(request,"postock.html",{"objects":ob_clients})
    